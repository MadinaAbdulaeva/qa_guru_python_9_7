import os
from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from webdriver_manager.core import archive


CURRENT_FILE = os.path.abspath(__file__)
CURRENT_DIR = os.path.dirname(CURRENT_FILE)
path = os.path.join(CURRENT_DIR, "tmp")
resources = os.path.join(CURRENT_DIR, "resources")
zip_path = os.path.join(resources, "test_archive.zip")

def test_pdf_file(create_directory, archive):
    with ZipFile(os.path.join(resources, 'test_archive.zip')) as archive:
        with archive.open('example.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            assert len(reader.pages) == 1
            assert "Пример PDF файла" in reader.pages[0].extract_text()
            file_stats = os.stat(os.path.join(path, 'example.pdf'))
            assert file_stats.st_size == archive.getinfo('example.pdf').file_size == 38329


def test_xlsx_file(create_directory, archive):
    with ZipFile(os.path.join(resources, 'test_archive.zip')) as archive:
        with archive.open('import_company.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert archive.getinfo('import_company.xlsx').file_size == 6809
            assert len(workbook.sheetnames) == 1
            assert sheet.max_row == 10
            assert sheet.max_column == 5
            assert sheet.cell(row=4, column=3).value == 'Отдел сервисного обслуживания'

def test_csv_file():
    with ZipFile(os.path.join(resources, 'test_archive.zip')) as archive:
        with archive.open('gross-domestic-product-june-2023-quarter.csv') as csv_file:
            reader = csv.DictReader(csv_file)
            print('Имена полей: ', reader.fieldnames)
            assert archive.getinfo('gross-domestic-product-june-2023-quarter.csv').file_size == 19034868




